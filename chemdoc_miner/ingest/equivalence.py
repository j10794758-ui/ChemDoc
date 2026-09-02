from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook

from chemdoc_miner.normalize import (
    canonical_company,
    canonical_grade,
    collapse_ws,
    looks_like_placeholder,
    normalize_grade,
    split_combo_grades,
)
from chemdoc_miner.paths import XLS_DIR

POWER_DREAM = "Power Dream"

_META_HEADERS = {
    "date",
    "other",
    "others",
    "igm analysis report",
    "chemistry group",
    "key",
    "key eq",
    "casnr",
    "cas",
    "product description",
    "photoinitiators",
    "product key",
    "product name",
    "additives name",
    "igm old name",
    "updated april.2.2026",
    "sample availability",
    "intial lead time",
    "bar preparation",
    "comment",
    "photoinitiator",
}


@dataclass
class EqMember:
    company: str
    grade: str
    grade_raw: str
    role: str = "product_code"


@dataclass
class EqGroup:
    eq_key: str | None
    domain: str
    chemistry: str | None
    cas: str | None
    chemistry_group: str | None
    source_file: str
    members: list[EqMember] = field(default_factory=list)

    @property
    def member_count(self) -> int:
        return len(self.members)

    @property
    def quality(self) -> str:
        if self.member_count >= 2:
            return "complete"
        if self.member_count == 1:
            return "sparse"
        return "placeholder"


def load_equivalence_groups() -> list[EqGroup]:
    groups: list[EqGroup] = []
    groups.extend(_offsets())
    groups.extend(_wide_sheet(XLS_DIR / "ECR Equivalence.xlsx", "ECR Equivalence", "ecr"))
    groups.extend(_wide_sheet(XLS_DIR / "ECR Equivalence.xlsx", "Amine", "ecr_amine"))
    groups.extend(_wide_sheet(XLS_DIR / "PI Equivalence.xlsx", "PI Equivalence", "pi"))
    groups.extend(_additives_sheet())
    return [g for g in groups if g.members]


def load_equivalence_rows() -> list[dict[str, Any]]:
    """Legacy pairwise export for backward compatibility."""
    rows: list[dict[str, Any]] = []
    for group in load_equivalence_groups():
        if group.member_count < 2:
            continue
        priority = {"offsets": 1, "ecr": 2, "ecr_amine": 2, "pi": 3, "additives": 4}.get(group.domain, 9)
        for i, m1 in enumerate(group.members):
            for m2 in group.members[i + 1 :]:
                if m1.company == m2.company and m1.grade == m2.grade:
                    continue
                rows.append(
                    _pair(
                        m1.company,
                        m1.grade,
                        m2.company,
                        m2.grade,
                        group.domain,
                        priority,
                        group.eq_key,
                        group.chemistry,
                        group.source_file,
                        group.cas,
                    )
                )
                rows.append(
                    _pair(
                        m2.company,
                        m2.grade,
                        m1.company,
                        m1.grade,
                        group.domain,
                        priority,
                        group.eq_key,
                        group.chemistry,
                        group.source_file,
                        group.cas,
                    )
                )
    return rows


def _offsets() -> list[EqGroup]:
    path = XLS_DIR / "Offsets Management 2026.xlsx"
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb["Competitive Landscape"]
    data = list(ws.iter_rows(values_only=True))
    wb.close()
    header = [collapse_ws(str(c)) if c else "" for c in data[0]]
    groups: list[EqGroup] = []
    for row_idx, row in enumerate(data[1:], start=2):
        if not row:
            continue
        pd_raw = row[3] if len(row) > 3 else None
        if looks_like_placeholder(pd_raw):
            continue
        pd_grades = split_combo_grades(str(pd_raw))
        if not pd_grades:
            continue
        members: list[EqMember] = []
        for pd in pd_grades:
            members.append(EqMember(POWER_DREAM, canonical_grade(pd, POWER_DREAM, domain="offsets"), str(pd_raw)))
        for col_index, company_header in enumerate(header):
            if col_index < 4 or col_index > 11:
                continue
            cell = row[col_index] if col_index < len(row) else None
            if looks_like_placeholder(cell):
                continue
            company = canonical_company(company_header)
            raw = str(cell)
            for grade in _split_cell(raw, company, domain="offsets"):
                members.append(EqMember(company, grade, raw))
        if not members:
            continue
        eq_key = f"OFFSET-{row_idx:03d}"
        groups.append(
            EqGroup(
                eq_key=eq_key,
                domain="offsets",
                chemistry=None,
                cas=None,
                chemistry_group=None,
                source_file=path.name,
                members=_dedupe_members(members),
            )
        )
    return groups


def _additives_sheet() -> list[EqGroup]:
    path = XLS_DIR / "Additives Equivalence.xlsx"
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]
    data = list(ws.iter_rows(values_only=True))
    wb.close()
    if not data:
        return []
    header = [collapse_ws(str(c)) if c else "" for c in data[0]]
    company_cols: list[tuple[int, str]] = []
    desc_col = cas_col = key_col = name_col = None
    for i, name in enumerate(header):
        low = name.lower().strip()
        if low == "additives name":
            name_col = i
            continue
        if low in {"key eq", "key"}:
            key_col = i
            continue
        if "product description" in low:
            desc_col = i
            continue
        if low.startswith("cas"):
            cas_col = i
            continue
        if not name or low in _META_HEADERS or low == "igm old name":
            continue
        company_cols.append((i, canonical_company(name)))
    groups: list[EqGroup] = []
    for row in data[1:]:
        if not row:
            continue
        eq_key = str(row[key_col]).strip() if key_col is not None and row[key_col] else None
        chemistry = collapse_ws(str(row[desc_col])) if desc_col is not None and row[desc_col] else None
        cas = str(row[cas_col]).strip() if cas_col is not None and row[cas_col] else None
        members: list[EqMember] = []
        if name_col is not None and row[name_col] and not looks_like_placeholder(row[name_col]):
            raw = str(row[name_col])
            for grade in _split_cell(raw, "IGM", domain="additives"):
                members.append(EqMember("IGM", grade, raw, role="product_code"))
        for col_index, company in company_cols:
            cell = row[col_index] if col_index < len(row) else None
            if looks_like_placeholder(cell):
                continue
            raw = str(cell)
            for grade in _split_cell(raw, company, domain="additives"):
                members.append(EqMember(company, grade, raw))
        members = _dedupe_members(members)
        if not members:
            continue
        groups.append(
            EqGroup(
                eq_key=eq_key,
                domain="additives",
                chemistry=chemistry or None,
                cas=cas or None,
                chemistry_group=None,
                source_file=path.name,
                members=members,
            )
        )
    return groups


def _wide_sheet(path: Path, sheet: str, domain: str) -> list[EqGroup]:
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb[sheet]
    data = list(ws.iter_rows(values_only=True))
    wb.close()
    if not data:
        return []
    header = [collapse_ws(str(c)) if c else "" for c in data[0]]
    company_cols: list[tuple[int, str]] = []
    desc_col = cas_col = key_col = chem_group_col = None
    for i, name in enumerate(header):
        low = name.lower().strip()
        if not name:
            continue
        if "product description" in low or low == "photoinitiators":
            desc_col = i
            continue
        if low.startswith("cas"):
            cas_col = i
            continue
        if low in {"key", "key eq"} or (low.startswith("key") and "eq" in low):
            key_col = i
            continue
        if "chemistry group" in low:
            chem_group_col = i
            continue
        if low in _META_HEADERS:
            continue
        company_cols.append((i, canonical_company(name)))
    groups: list[EqGroup] = []
    for row in data[1:]:
        if not row:
            continue
        eq_key = str(row[key_col]).strip() if key_col is not None and row[key_col] else None
        chemistry = collapse_ws(str(row[desc_col])) if desc_col is not None and row[desc_col] else None
        cas = str(row[cas_col]).strip() if cas_col is not None and row[cas_col] else None
        chemistry_group = (
            collapse_ws(str(row[chem_group_col])) if chem_group_col is not None and row[chem_group_col] else None
        )
        members: list[EqMember] = []
        for col_index, company in company_cols:
            cell = row[col_index] if col_index < len(row) else None
            if looks_like_placeholder(cell):
                continue
            raw = str(cell)
            for grade in _split_cell(raw, company, domain=domain):
                members.append(EqMember(company, grade, raw))
        members = _dedupe_members(members)
        if not members:
            continue
        groups.append(
            EqGroup(
                eq_key=eq_key,
                domain=domain,
                chemistry=chemistry or None,
                cas=cas or None,
                chemistry_group=chemistry_group or None,
                source_file=path.name,
                members=members,
            )
        )
    return groups


def _dedupe_members(members: list[EqMember]) -> list[EqMember]:
    seen: set[tuple[str, str]] = set()
    out: list[EqMember] = []
    for m in members:
        key = (m.company, m.grade)
        if key in seen:
            continue
        seen.add(key)
        out.append(m)
    return out


def _split_cell(value: str, company: str = "", *, domain: str | None = None) -> list[str]:
    text = collapse_ws(value)
    text = re.sub(r"\((close|better property|offset ink)\)", "", text, flags=re.I)
    name = canonical_company(company) if company else ""

    if name == "IGM" and domain == "pi":
        grades: list[str] = []
        for pat in (
            r"(?i)(?:omnirad|esacure|omnipol|photomer|ph)\s*[-]?\s*(\d+[a-z0-9\-]*)",
            r"(?i)\b(\d{3,4}[a-z]?)\b",
        ):
            for m in re.finditer(pat, text):
                g = canonical_grade(m.group(1), name, domain=domain)
                if g and g not in grades:
                    grades.append(g)
        if grades:
            return grades

    parts = re.split(r"[,;/]|\band\b", text)
    grades = []
    for part in parts:
        part = collapse_ws(part)
        if looks_like_placeholder(part):
            continue
        if len(part) > 80:
            continue
        grade = canonical_grade(part, name, domain=domain) if name else normalize_grade(part)
        if grade and grade not in grades:
            grades.append(grade)
    return grades


def _pair(
    src_company: str,
    src_grade: str,
    dst_company: str,
    dst_grade: str,
    eq_kind: str,
    priority: int,
    eq_key: str | None,
    chemistry: str | None,
    source_file: str,
    cas: str | None = None,
) -> dict[str, Any]:
    return {
        "src_company": src_company,
        "src_grade": src_grade,
        "dst_company": dst_company,
        "dst_grade": dst_grade,
        "eq_key": eq_key,
        "chemistry": chemistry,
        "cas": cas,
        "source_file": source_file,
        "eq_kind": eq_kind,
        "priority": priority,
        "needs_review": 0,
    }


def iter_alias_hints(groups: Iterator[EqGroup] | list[EqGroup]) -> list[tuple[str, str, str]]:
    from chemdoc_miner.normalize import grade_query_variants

    aliases: list[tuple[str, str, str]] = []
    seen: set[tuple[str, str, str]] = set()
    for group in groups:
        for member in group.members:
            company, grade = member.company, member.grade
            for alias in grade_query_variants(member.grade_raw, company, domain=group.domain):
                key = (alias, company, grade)
                if key not in seen:
                    seen.add(key)
                    aliases.append(key)
            for alias in grade_query_variants(grade, company, domain=group.domain):
                key = (alias, company, grade)
                if key not in seen:
                    seen.add(key)
                    aliases.append(key)
            key = (grade, company, grade)
            if key not in seen:
                seen.add(key)
                aliases.append(key)
    return aliases
